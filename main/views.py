import matplotlib.pyplot as plt
import os
import openpyxl
import re
import pandas as pd
import matplotlib

matplotlib.use('Agg')

from datetime import datetime
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
from django.shortcuts import render, redirect
from .models import Task
from .forms import TaskForm, FileForm, TaskUpForm, SearchForm
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView  # подробное представление задачи
from django.views.generic.edit import CreateView, UpdateView, DeleteView, FormView
from django.urls import reverse_lazy
from django.contrib.auth.views import LoginView
from django.contrib.auth.mixins import LoginRequiredMixin  # для ограничения доступа на страницу
from django.contrib.auth.forms import UserCreationForm  # автоматическая форма создания пользователя
from django.contrib.auth import login  # при регистрации автоматически войдет в систему
from django.http import FileResponse, Http404


class CustomLoginView(LoginView):
    template_name = 'main/login.html'
    fields = '__all__'
    redirect_authenticated_user = True  # неаутентифицированные пользователи не смогут зайти на страницу

    def get_success_url(self):
        return reverse_lazy('tasks')


class RegisterPage(FormView):
    template_name = 'main/register.html'
    form_class = UserCreationForm
    redirect_authenticated_user = True
    success_url = reverse_lazy('tasks')

    def form_valid(self, form):
        user = form.save()
        if user is not None:
            login(self.request, user)
        return super(RegisterPage, self).form_valid(form)

    def get(self, *args, **kwargs):  # функция для перенаправления авторизованного пользователя на список задач
        if self.request.user.is_authenticated:
            return redirect('tasks')
        return super(RegisterPage, self).get(*args, **kwargs)


class TaskList(LoginRequiredMixin, ListView):  # ищет шaблон с профиксом _list: main/task_list.html + ограничения на доступ -> перенаправление на страницу login (см settings.py LOGIN_URL)
    model = Task
    context_object_name = 'tasks'  # изменение названия объекта TaskList

    def get_context_data(self, **kwargs):  # для настройки вывода списка задач только зарегистрированного пользователя
        context = super().get_context_data(**kwargs)
        if self.request.user.is_superuser:
            context['tasks'] = context['tasks'].all()
            context['count'] = context['tasks'].filter(is_complete=False).count()
        else:
            context['tasks'] = context['tasks'].filter(Q(user=self.request.user) | Q(author=self.request.user))
            context['count'] = context['tasks'].filter(is_complete=False).count()

        search_input = self.request.GET.get('search-area') or ''
        if search_input:
            context['tasks'] = context['tasks'].filter(task__iregex=search_input)

        context['search_input'] = search_input

        return context


class TaskDetail(LoginRequiredMixin, DetailView):  # ищет шaблон с профиксом _detail: main/task_detail.html + ограничения на доступ -> перенаправление на страницу login (см settings.py LOGIN_URL)
    model = Task
    context_object_name = 'task'


class TaskCreate(LoginRequiredMixin, CreateView):  # ищет шaблон с профиксом _form. автоматически создает поля модели
    model = Task
    form_class = TaskForm
    success_url = reverse_lazy('tasks')  # ленивый реверс, перенаправления на страницу

    def form_valid(self, form):  # для автоматического заполнения поля author и user , метод есть в классе CreateView
        if self.request.user:
            form.instance.author = self.request.user
        if form.instance.user is None:
            form.instance.user = self.request.user
        return super(TaskCreate, self).form_valid(form)


class TaskUpdate(LoginRequiredMixin, UpdateView):  # ищет шaблон с профиксом _form. автоматически создает поля модели
    template_name = 'main/task_form_update.html'
    model = Task
    form_class = TaskUpForm
    success_url = reverse_lazy('tasks')  # ленивый реверс, перенаправления на страницу


def update(request, task_id): # Для реализации галочки "Выполнено"
    task = Task.objects.get(id=task_id)
    task.is_complete = not task.is_complete
    task.save()
    return redirect('tasks')


class TaskDelete(LoginRequiredMixin, DeleteView):  # ищет шaблон с префиксом _confirm_delete
    model = Task
    context_object_name = 'task'
    success_url = reverse_lazy('tasks')  # ленивый реверс, перенаправления на страницу после удаления задачи


def index(request):
    context = {'title': 'Информация'}
    return render(request, 'main/index.html', context)


def search(request):
    result_dict = None  # Инициализация переменной для хранения результата
    form_search = SearchForm()
    form_upload = FileForm()  # Форма для загрузки файла
    upload_success = False  # Переменная для сообщения о загрузке файла
    is_admin = request.user.is_superuser or request.user.groups.filter(name='Администратор').exists()

    if request.method == 'POST':
        if 'upload' in request.POST:  # Проверяем, была ли отправлена форма загрузки
            form_upload = FileForm(request.POST, request.FILES)
            if form_upload.is_valid():
                uploaded_file = request.FILES['file']
                fs = FileSystemStorage()
                uploaded_file_path = os.path.join(fs.location, 'SVOD.xlsx')

                # Сохраняем файл на сервере
                with open(uploaded_file_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)
                upload_success = True  # Устанавливаем переменную, чтобы отобразить сообщение

        else:  # Обработка формы поиска
            form_search = SearchForm(request.POST)
            if form_search.is_valid():  # Проверяем валидность формы
                ppz_number = form_search.cleaned_data['ppz_number']
                try:
                    df = pd.read_excel('media/SVOD.xlsx')  # Путь к загруженному файлу
                    result = df[df['Номер ПЗ'] == ppz_number]  # Предполагаем, что 'Номер ПЗ' - это название столбца
                    if not result.empty:
                        result_dict = result.iloc[0].to_dict()  # Получаем первую найденную запись как словарь
                        result_dict = {key.replace(' ', '_'): value for key, value in result_dict.items()}
                    else:
                        result_dict = {'error': 'Запись не найдена.'}  # Если запись не найдена
                except FileNotFoundError:
                    result_dict = {'error': 'Файл не найден. Убедитесь, что он загружен администратором.'}
                except Exception as e:
                    result_dict = {'error': str(e)}  # Обработка других исключений

    context = {
        'title': 'Поиск закупки',
        'form2': form_search,  # Форма для поиска
        'form_upload': form_upload,  # Форма для загрузки файла
        'result': result_dict,  # Результат поиска
        'upload_success': upload_success,  # Передаем переменную для сообщения
        'is_admin': is_admin,  # Передаем переменную в контекст
    }
    return render(request, 'main/search.html', context)


def upload_file(request):
    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES)

        if form.is_valid():
            file = form.cleaned_data['file']
            fs = FileSystemStorage()

            uploaded_file_path = os.path.join(fs.location, 'Аnalis.xlsx')

            with open(uploaded_file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)

            flag = False

            # Обработка загруженного файла
            wb = openpyxl.load_workbook(uploaded_file_path)  # Используем полный путь к файлу
            sheet = wb.active

            start_time = datetime.now()  # время начала выполнения
            # print(start_time)

            # Получаем наименования столбцов
            headers = {cell.value: cell.column for cell in sheet[1]}  # Словарь с наименованиями столбцов и их номерами

            # Используем наименование столбца вместо номера
            column_name = 'Комментарий для статуса'  # Замените на фактическое наименование столбца
            column_index = headers.get(column_name)
            max_column = sheet.max_column

            # Название для предпоследнего столбца
            sheet.cell(row=1, column=max_column + 1, value='Количество доработок')
            sheet.cell(row=1, column=max_column + 2, value='Доработки подробно')

            if column_index is None:
                return render(request, 'main/analysis.html',
                              {'title': 'Анализ', 'form': form, 'error': 'Загружена некорректная форма!'})

            for i in range(1, sheet.max_row):
                data = sheet.cell(row=i + 1, column=column_index).value
                if data is not None:
                    data = re.sub(r'\«((?:.|\n)*?)\»', lambda x: x.group(0).replace('\n', ' '), data, flags=re.DOTALL)
                    lines = data.split("\n")
                    formatted_data = []
                    dates = []
                for line in lines:
                    if line.strip():
                        end_index = line.find('»')
                        first_phrase = line[:end_index + 1]
                        line = line[end_index + 1:].strip()
                        parts = line.split()
                        if len(parts) > 1:
                            formatted_data.append({
                                "Статус": first_phrase,
                                "Дата": parts[0],
                                "Время": parts[1],
                                "Комментарий": " ".join(parts[3:])
                            })
                            dates.append((first_phrase, parts[0]))

                dorabotka_date = None
                back_date = None
                count = 0
                count_back = set()
                res = ''

                for j in range(len(dates)):
                    if dorabotka_date and back_date:
                        dorabotka_date = None
                        back_date = None

                    if dates[j][0] == '«Доработка Заказчиком Доработка ППЗ»' or dates[j][0] == '«Доработка»':
                        dorabotka_date = dates[j][1]
                        count += 1

                    elif (dates[j][0] == '«Анализ цены Д647 Назначение исполнителя»' or dates[j][0]
                          == '«Формирование ППЗ Заказчиком Согласование тендерного подразделения»') and dorabotka_date:
                        back_date = dates[j][1]

                        if dorabotka_date and back_date:
                            dorabotka_date = datetime.strptime(dorabotka_date, '%d.%m.%Y')
                            back_date = datetime.strptime(back_date, '%d.%m.%Y')
                            difference = back_date - dorabotka_date
                            res += (f'Доработка №{count}: {difference.days} дн. Месяц направления на доработку - '
                                    f'{dorabotka_date.month}, месяц отработки - {back_date.month};\n')

                if back_date == None and dorabotka_date:
                    if sheet.cell(row=i + 1, column=7).value != 'ППЗ утверждена':
                        now = datetime.now()
                        difference = now - datetime.strptime(dorabotka_date, '%d.%m.%Y')
                        res += f'Доработка №{count}: не отработано с {dorabotka_date} ({difference.days} дн.);\n'

                    elif sheet.cell(row=i + 1, column=7).value == 'ППЗ утверждена' or sheet.cell(row=i + 1,
                                                                                                 column=7).value == 'ППЗ аннулирована':
                        now = datetime.now()  # подтянуть дату статуса
                        difference = now - datetime.strptime(dorabotka_date, '%d.%m.%Y')
                        res += f'Доработка №{count}: {difference.days} дн. Месяц направления на доработку - {dorabotka_date}, месяц отработки - {now.month};\n'

                elif count == 0:
                    res = 'Доработок нет'

                # считаем количество доработок всего
                count_back.add(count)
                cell2 = sheet.cell(row=i + 1, column=max_column + 1)
                cell2.value = max(count_back)

                # выводим информацию о доработках
                cell = sheet.cell(row=i + 1, column=max_column + 2)
                cell.value = res

        # Сохраняем результат
        result_file_path = os.path.join(fs.location, 'Result.xlsx')
        wb.save(result_file_path)

        flag = True

        end_time = datetime.now()  # время окончания выполнения
        execution_time = end_time - start_time
        quantity = sheet.max_row - 1

        # Вернуть результат
        context = {'title': 'Анализ',
                   'form': form,
                   'result': formatted_data,
                   'flag': flag,
                   'time': execution_time,
                   'quantity': quantity,
                   }
        return render(request, 'main/analysis.html', context)

    else:
        form = FileForm()

    # Читаем данные из Excel
    data = pd.read_excel(os.path.join(FileSystemStorage().location, 'Result.xlsx'))

    # Генерация графиков
    generate_charts(data)

    context = {
        'title': 'Анализ',
        'form': form,
        'data': data,
    }
    return render(request, 'main/analysis.html', context)


def download_file(request):
    # Путь к файлу, который нужно скачать
    file_path = os.path.join('media/', 'Result.xlsx')

    # Проверяем, существует ли файл
    if os.path.exists(file_path):
        # Возвращаем файл как ответ
        response = FileResponse(open(file_path, 'rb'),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Result.xlsx"'
        return response
    else:
        # Если файл не найден, возвращаем 404 ошибку
        raise Http404("Файл не найден")


def generate_charts(data):
    # Генерация круговой диаграммы доля закупок по способу(количественное выражение)
    plt.figure(figsize=(10, 6))
    data['Способ закупки'].value_counts().plot.pie(autopct='%1.0f%%', fontsize=14, labels=None)
    plt.legend(labels=data['Способ закупки'].value_counts().index,
               loc="upper left", bbox_to_anchor=(1, 1), fontsize=12)
    plt.title('по количеству', fontsize=16)
    plt.ylabel('')
    plt.tight_layout()
    plt.savefig('media/pie_chart.png')
    plt.close()

    # Генерация круговой диаграммы доля закупок по способу(суммарное выражение)
    plt.figure(figsize=(10, 6))
    procurement_summary = data.groupby('Способ закупки')['Плановая цена закупки(без НДС)'].sum()
    procurement_summary.plot.pie(autopct='%1.0f%%', fontsize=14, labels=None)
    plt.legend(labels=data.groupby('Способ закупки')['Плановая цена закупки(без НДС)'].sum().index,
               loc="upper left", bbox_to_anchor=(1, 1), fontsize=12)
    plt.title('по планируемой сумме', fontsize=16)
    plt.ylabel('')  # Убираем метку Y
    plt.tight_layout()
    plt.savefig('media/pie_chart_by_procurement.png')
    plt.close()

    # Генерация круговой диаграммы доля закупок по статусу(количественное выражение)
    plt.figure(figsize=(10, 6))
    data['Статус закупки'].value_counts().plot.pie(autopct='%1.0f%%', fontsize=14, labels=None)
    plt.legend(labels=data['Статус закупки'].value_counts().index,
               loc="upper left", bbox_to_anchor=(1, 1), fontsize=12)
    plt.title('по количеству', fontsize=16)
    plt.ylabel('')
    plt.tight_layout()
    plt.savefig('media/pie_chart_status.png')
    plt.close()

    # Генерация круговой диаграммы доля закупок по статусу(суммарное выражение)
    plt.figure(figsize=(10, 6))
    procurement_summary = data.groupby('Статус закупки')['Плановая цена закупки(без НДС)'].sum()
    procurement_summary.plot.pie(autopct='%1.0f%%', fontsize=14, labels=None)
    plt.legend(labels=data.groupby('Статус закупки')['Плановая цена закупки(без НДС)'].sum().index,
               loc="upper left", bbox_to_anchor=(1, 1), fontsize=12)
    plt.title('по планируемой сумме', fontsize=16)
    plt.ylabel('')
    plt.tight_layout()
    plt.savefig('media/pie_chart_status_by_procurement.png')
    plt.close()

    # Генерация точечной диаграммы
    grouped_data = data.groupby('Наименование филиала')['Количество доработок'].sum().reset_index()
    grouped_data2 = data.groupby('Наименование филиала')['Количество доработок'].count().reset_index()

    plt.figure(figsize=(8, 4))
    plt.scatter(grouped_data['Наименование филиала'],
                grouped_data['Количество доработок'],
                color='blue',
                s=50,
                label='Количество доработок')
    plt.scatter(grouped_data2['Наименование филиала'],
                grouped_data2['Количество доработок'],
                color='red',
                s=50,
                label='Всего закупок')
    plt.tick_params(axis='both', labelsize=8)
    plt.legend(loc="best", fontsize=8)
    # Настройки графика
    plt.title('', fontsize=8)
    plt.xlabel('', fontsize=8)
    plt.ylabel('', fontsize=8)
    plt.grid(True)

    # Добавляем аннотации для точек
    for i in range(grouped_data.shape[0]):
        plt.annotate(f"{grouped_data['Количество доработок'][i]}",
                     (grouped_data['Наименование филиала'][i], grouped_data['Количество доработок'][i]),
                     textcoords="offset points",
                     xytext=(0, 10),
                     fontsize=8,
                     ha='right',
                     va="top")

    for i in range(grouped_data2.shape[0]):
        plt.annotate(f"{grouped_data2['Количество доработок'][i]}",
                     (grouped_data2['Наименование филиала'][i], grouped_data2['Количество доработок'][i]),
                     textcoords="offset points",
                     xytext=(0, 10),
                     fontsize=8,
                     ha='left',
                     va="bottom")
    # Сохраняем график
    plt.tight_layout()
    plt.savefig('media/scatter_chart.png')
    plt.close()


def about(request):
    context = {'title': 'О проекте'}
    return render(request, 'main/about.html', context)
